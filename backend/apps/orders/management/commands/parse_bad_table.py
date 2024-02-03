from apps.customers.models import Customer
from django.core.management.base import BaseCommand
from django.db import transaction
from openpyxl import load_workbook

from apps.utils import Formatter

# Признаки для определения покупателя как "мастер"
MASTER_DATA = [
    "100",
    "150",
    "180",
    "250",
    "300",
    "350",
    "мастер",
    "видео",
    "марафон",
    "курс",
    "Я сама",
    "помоги",
    "Мастер-класс",
    "Сама",
    "Видеокурс",
    "Видеоурок",
    "ПОМОГИ",
    "Курс",
    "Марафон"
]

class Command(BaseCommand, Formatter):
    help = "Парсит древние таблицы с заказами и записывает их в существующую БД"

    def __init__(self):
        super(Command, self).__init__()
        self.master = {}
        self.default_customer = {}

    def handle(self, *args, **kwargs):
        # Загружаем Excel
        workbook = load_workbook(filename='all_data.xlsx')
        sheet = workbook.active

        # Проходим по строкам начиная со второй (исключая заголовки)
        for row in sheet.iter_rows(min_row=2):
            product = row[4].value  # столбец E - products
            email = row[1].value  # столбец B - Email
            name = row[0].value  # столбец A - Name
            price = row[5].value  # столбец F - price
            phone = row[13].value  # столбец N - Phone

            if all([product, price, email]):
                # Проверяем наличие ключевых слов в продукте
                if any(keyword in product for keyword in MASTER_DATA):
                    if email in self.master:
                        self.master[email][2] += round(float(price))
                    else:
                        self.master[email] = [name, phone, round(float(price))]

                # Отдельный словарь с инфой о простых покупателях
                else:
                    if email in self.default_customer:
                        self.default_customer[email][2] += round(float(price))
                    else:
                        self.default_customer[email] = [name, phone, round(float(price))]

        if self.master:
            self.save_customers()

    @transaction.atomic
    def save_customers(self):
        """Сохранение клиентов в БД"""

        for email, values in self.master.items():
            last_name, first_name, patronymic_name = self.format_full_name(values[0]) if values else ""
            phone = self.format_phone(values[1])
            sum_orders = values[2] if values[2] else 0

            try:
                Customer.objects.get(email=email)
            except Customer.DoesNotExist:
                Customer.objects.create(
                    first_name=first_name,
                    last_name=last_name,
                    patronymic_name=patronymic_name,
                    email=email,
                    phone_number=phone,
                    sum_old_orders=sum_orders,
                )
